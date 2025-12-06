
from __future__ import annotations
from typing import Iterable, List, Dict, Any
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def courses_to_xlsx_bytes(rows: List[Dict[str, Any]], sheet_name: str = "Courses") -> bytes:
    """
    rows: list of dict, each dict is a row
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    if not rows:
        ws.append(["No data"])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(rows[0].keys())
    ws.append(headers)

    # header style
    header_font = Font(bold=True)
    for col_idx, _h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # data
    for r in rows:
        ws.append([r.get(h) for h in headers])

    # autosize columns
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_filename(prefix: str = "courses") -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.xlsx"
